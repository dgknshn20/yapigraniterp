import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from django.http import HttpResponse
from datetime import datetime


def export_transactions_to_excel(queryset):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Kasa Hareketleri"

    headers = [
        "ID",
        "Tarih",
        "İşlem Tipi",
        "Açıklama",
        "Tutar",
        "Kaynak Hesap",
        "Hedef Hesap",
        "İlgili Cari",
        "Proje/Sözleşme",
    ]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_num, txn in enumerate(queryset, 2):
        worksheet.cell(row=row_num, column=1, value=txn.id)
        worksheet.cell(row=row_num, column=2, value=txn.date)
        worksheet.cell(row=row_num, column=3, value=txn.get_transaction_type_display())
        worksheet.cell(row=row_num, column=4, value=txn.description)
        worksheet.cell(row=row_num, column=5, value=float(txn.amount))
        worksheet.cell(row=row_num, column=6, value=txn.source_account.name if txn.source_account else "-")
        worksheet.cell(row=row_num, column=7, value=txn.target_account.name if txn.target_account else "-")
        worksheet.cell(row=row_num, column=8, value=txn.related_customer.name if txn.related_customer else "")
        worksheet.cell(
            row=row_num,
            column=9,
            value=txn.related_contract.project_name if txn.related_contract else "",
        )

    for col in worksheet.columns:
        max_length = 0
        column_letter = col[0].column_letter
        for cell in col:
            try:
                cell_length = len(str(cell.value)) if cell.value is not None else 0
                if cell_length > max_length:
                    max_length = cell_length
            except Exception:
                pass
        worksheet.column_dimensions[column_letter].width = max_length + 2

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f"Kasa_Hareketleri_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    workbook.save(response)
    return response
